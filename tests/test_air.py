import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from air.inputs import InputError, read_queries
from air.remote_worker import RemoteWorker
from air.worker import main as run_legacy_worker
from app import create_app


def uploaded_file(files, field):
    entries = files.items() if isinstance(files, dict) else files
    return next(value for key, value in entries if key == field)


class AirTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        database = Path(self.temp.name) / "air-test.db"
        self.app = create_app({
            "TESTING": True,
            "DATABASE_URL": f"sqlite:///{database.as_posix()}",
            "AIR_WORKER_TOKEN": "test-worker-secret",
        })
        self.client = self.app.test_client()
        self.client_id = "a" * 32
        self.client_headers = {"X-AIR-Client-ID": self.client_id}
        self.worker_headers = {"Authorization": "Bearer test-worker-secret"}

    def tearDown(self):
        self.app.extensions["air_store"].engine.dispose()
        self.temp.cleanup()

    def upload(self, text=b"first query\nsecond query", filename="queries.txt"):
        return self.client.post(
            "/jobs",
            data={"file": (io.BytesIO(text), filename)},
            content_type="multipart/form-data",
            headers=self.client_headers,
        )

    def test_brand_queue_interface_and_limit(self):
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertIn(b"AMEX AIR", response.data)
        self.assertIn(b"AI Insights &amp; Responses", response.data)
        self.assertIn(b"35 queries", response.data)
        self.assertIn(b"Request status", response.data)

    def test_supported_inputs_and_upload_limit(self):
        self.assertEqual(["one", "two"], read_queries("q.txt", b"one\ntwo\none\n"))
        self.assertEqual(["one", "two"], read_queries("q.csv", b"Prompt,Other\none,x\ntwo,y\n"))
        book = Workbook()
        sheet = book.active
        sheet.append(["Prompt"])
        sheet.append(["one"])
        sheet.append(["two"])
        stream = io.BytesIO()
        book.save(stream)
        self.assertEqual(["one", "two"], read_queries("q.xlsx", stream.getvalue()))
        with self.assertRaisesRegex(InputError, "at most 500"):
            read_queries("q.txt", "\n".join(f"query {i}" for i in range(501)).encode())

    def test_upload_is_durable_and_visible_to_same_browser(self):
        response = self.upload()
        self.assertEqual(202, response.status_code, response.get_data(as_text=True))
        job = response.get_json()["job"]
        self.assertEqual("queued", job["state"])
        self.assertEqual(2, job["total"])
        returning_browser = self.app.test_client()
        listing = returning_browser.get("/jobs", headers=self.client_headers).get_json()
        self.assertEqual(job["run_id"], listing["jobs"][0]["run_id"])
        self.assertFalse(listing["worker"]["online"])
        other = returning_browser.get("/jobs", headers={"X-AIR-Client-ID": "b" * 32}).get_json()
        self.assertEqual([], other["jobs"])

    def test_worker_claim_progress_complete_and_download(self):
        run_id = self.upload().get_json()["job"]["run_id"]
        denied = self.client.post("/worker/claim", json={"worker_id": "laptop"})
        self.assertEqual(401, denied.status_code)
        claim = self.client.post("/worker/claim", json={"worker_id": "laptop"}, headers=self.worker_headers)
        self.assertEqual(run_id, claim.get_json()["job"]["run_id"])
        progress = self.client.post(
            f"/worker/jobs/{run_id}/progress",
            json={"worker_id": "laptop", "completed": 1, "current_query": "second query", "message": "One done."},
            headers=self.worker_headers,
        )
        self.assertFalse(progress.get_json()["cancel_requested"])
        completed = self.client.post(
            f"/worker/jobs/{run_id}/complete",
            data={
                "worker_id": "laptop",
                "success_count": "1",
                "failed_count": "1",
                "workbook": (io.BytesIO(b"test-workbook"), "results.xlsx"),
            },
            content_type="multipart/form-data",
            headers=self.worker_headers,
        )
        self.assertEqual(200, completed.status_code)
        job = self.client.get(f"/jobs/{run_id}", headers=self.client_headers).get_json()
        self.assertEqual("completed", job["state"])
        self.assertEqual(2, job["completed"])
        download = self.client.get(job["download_url"])
        self.assertEqual(b"test-workbook", download.data)
        self.assertEqual(404, self.client.get(f"/jobs/{run_id}/download?token=wrong").status_code)

    def test_cancel_queued_and_processing_jobs(self):
        queued_id = self.upload(b"queued").get_json()["job"]["run_id"]
        response = self.client.post(f"/jobs/{queued_id}/cancel", headers=self.client_headers)
        self.assertEqual(200, response.status_code, response.get_data(as_text=True))
        queued = response.get_json()["job"]
        self.assertEqual("cancelled", queued["state"])
        self.assertIn("download_url", queued)
        empty_output = self.client.get(queued["download_url"])
        self.assertEqual(200, empty_output.status_code)
        empty_book = load_workbook(io.BytesIO(empty_output.data), read_only=True, data_only=True)
        self.assertEqual(1, empty_book["Responses"].max_row)
        empty_book.close()
        processing_id = self.upload(b"processing").get_json()["job"]["run_id"]
        self.client.post("/worker/claim", json={"worker_id": "laptop"}, headers=self.worker_headers)
        self.client.post(
            f"/worker/jobs/{processing_id}/progress",
            json={"worker_id": "laptop", "completed": 1, "current_query": "processing", "message": "One done."},
            headers=self.worker_headers,
        )
        processing = self.client.post(f"/jobs/{processing_id}/cancel", headers=self.client_headers).get_json()["job"]
        self.assertTrue(processing["cancel_requested"])
        partial = self.client.post(
            f"/worker/jobs/{processing_id}/complete",
            data={
                "worker_id": "laptop",
                "success_count": "1",
                "failed_count": "0",
                "workbook": (io.BytesIO(b"partial-workbook"), "results.xlsx"),
            },
            content_type="multipart/form-data",
            headers=self.worker_headers,
        )
        self.assertEqual("cancelled", partial.get_json()["state"])
        job = self.client.get(f"/jobs/{processing_id}", headers=self.client_headers).get_json()
        self.assertEqual("cancelled", job["state"])
        self.assertEqual(1, job["completed"])
        self.assertIn("Partial results", job["message"])
        self.assertEqual(b"partial-workbook", self.client.get(job["download_url"]).data)

    def test_laptop_worker_builds_and_uploads_workbook(self):
        from unittest.mock import MagicMock, patch

        class FakeCollector:
            def collect(self, query):
                return {"status": "Success", "response": f"answer: {query}", "parsed_json": "{}", "execution_time": 0.1}

        worker = RemoteWorker("https://example.test", "secret", worker_id="laptop")
        progress = []
        uploads = []
        worker.update = lambda run_id, completed, current_query, message: progress.append((completed, current_query)) or False

        def capture(path, **kwargs):
            workbook = uploaded_file(kwargs["files"], "workbook")
            batches = [value for key, value in kwargs["files"] if key == "batches"]
            uploads.append((path, kwargs["data"], workbook[1].read(), batches))
            return MagicMock()

        worker.post = capture
        with patch("air.remote_worker.GoogleAIOverviewCollector", FakeCollector), patch.dict("os.environ", {"AIR_QUERY_DELAY_SECONDS": "0"}):
            worker.process({"run_id": "c" * 32, "filename": "queries.txt", "queries": ["one", "two"]})
        self.assertEqual([0, 1, 1, 2], [item[0] for item in progress])
        self.assertEqual("/worker/jobs/cccccccccccccccccccccccccccccccc/complete", uploads[0][0])
        self.assertEqual(["amex_air_batch_1.xlsx"], [batch[0] for batch in uploads[0][3]])
        workbook = load_workbook(io.BytesIO(uploads[0][2]), read_only=True, data_only=True)
        self.assertEqual("answer: one", workbook["Responses"]["C2"].value)
        workbook.close()

    def test_laptop_worker_uploads_partial_workbook_when_cancelled(self):
        from unittest.mock import MagicMock, patch

        collected = []

        class FakeCollector:
            def collect(self, query):
                collected.append(query)
                return {"status": "Success", "response": f"answer: {query}", "parsed_json": "{}", "execution_time": 0.1}

        worker = RemoteWorker("https://example.test", "secret", worker_id="laptop")
        uploads = []
        worker.update = lambda _run_id, completed, _query, _message, **_kwargs: completed == 1

        def capture(path, **kwargs):
            workbook = uploaded_file(kwargs["files"], "workbook")
            uploads.append((path, kwargs["data"], workbook[1].read()))
            return MagicMock()

        worker.post = capture
        with patch("air.remote_worker.GoogleAIOverviewCollector", FakeCollector), patch.dict("os.environ", {"AIR_QUERY_DELAY_SECONDS": "0"}):
            worker.process({"run_id": "p" * 32, "filename": "queries.txt", "queries": ["one", "two", "three"]})
        self.assertEqual(["one"], collected)
        self.assertEqual("/worker/jobs/pppppppppppppppppppppppppppppppp/complete", uploads[0][0])
        self.assertEqual("1", str(uploads[0][1]["success_count"]))
        workbook = load_workbook(io.BytesIO(uploads[0][2]), read_only=True, data_only=True)
        self.assertEqual(2, workbook["Responses"].max_row)
        self.assertEqual("one", workbook["Responses"]["A2"].value)
        workbook.close()

    def test_laptop_worker_cools_down_and_retries_same_query_after_captcha(self):
        from unittest.mock import MagicMock, patch

        collected = []

        class FakeCollector:
            def collect(self, query):
                collected.append(query)
                if len(collected) == 1:
                    return {
                        "status": "Failed",
                        "response": "Google displayed a CAPTCHA or unusual-traffic block.",
                        "parsed_json": "",
                        "execution_time": 0.1,
                    }
                return {"status": "Success", "response": f"answer: {query}", "parsed_json": "{}", "execution_time": 0.1}

        worker = RemoteWorker("https://example.test", "secret", worker_id="laptop")
        updates = []
        uploads = []
        worker.update = lambda _run_id, completed, query, message, **_kwargs: updates.append((completed, query, message)) or False

        def capture(path, **kwargs):
            workbook = uploaded_file(kwargs["files"], "workbook")
            uploads.append((path, kwargs["data"], workbook[1].read()))
            return MagicMock()

        worker.post = capture
        settings = {
            "AIR_QUERY_DELAY_SECONDS": "0",
            "AIR_CAPTCHA_COOLDOWN_SECONDS": "0",
            "AIR_CAPTCHA_RETRIES": "1",
            "AIR_BATCH_REST_EVERY": "0",
        }
        with patch("air.remote_worker.GoogleAIOverviewCollector", FakeCollector), patch.dict("os.environ", settings):
            worker.process({"run_id": "r" * 32, "filename": "queries.txt", "queries": ["one"]})

        self.assertEqual(["one", "one"], collected)
        self.assertTrue(any("temporarily blocked" in message for _, _, message in updates))
        self.assertEqual("1", str(uploads[0][1]["success_count"]))
        self.assertEqual("0", str(uploads[0][1]["failed_count"]))

    def test_36_queries_create_two_batches_and_one_combined_workbook(self):
        from unittest.mock import MagicMock, patch

        class FakeCollector:
            def collect(self, query):
                return {
                    "status": "Success", "response": query, "parsed_json": "{}",
                    "top_blue_links": "[]", "execution_time": 0.1,
                }

        worker = RemoteWorker("https://example.test", "secret", worker_id="laptop")
        worker.update = lambda *_args, **_kwargs: False
        uploads = []

        def capture(path, **kwargs):
            uploads.append((path, kwargs))
            return MagicMock()

        worker.post = capture
        settings = {
            "AIR_QUERY_DELAY_SECONDS": "0", "AIR_BATCH_REST_EVERY": "0",
            "AIR_INTER_BATCH_COOLDOWN_SECONDS": "0",
        }
        with patch("air.remote_worker.GoogleAIOverviewCollector", FakeCollector), patch.dict("os.environ", settings):
            worker.process({"run_id": "b" * 32, "filename": "queries.txt", "queries": [f"q{i}" for i in range(36)]})

        files = uploads[0][1]["files"]
        batches = [value for key, value in files if key == "batches"]
        self.assertEqual(["amex_air_batch_1.xlsx", "amex_air_batch_2.xlsx"], [item[0] for item in batches])
        combined = uploaded_file(files, "workbook")
        book = load_workbook(io.BytesIO(combined[1].read()), read_only=True, data_only=True)
        self.assertEqual(37, book["Responses"].max_row)
        book.close()

    def test_persistent_captcha_stops_after_two_retries_with_partial_output(self):
        from unittest.mock import MagicMock, patch

        attempts = []

        class BlockedCollector:
            def collect(self, query):
                attempts.append(query)
                return {
                    "status": "Failed",
                    "response": "Google displayed a CAPTCHA or unusual-traffic block.",
                    "parsed_json": "", "top_blue_links": "[]", "execution_time": 0.1,
                }

        worker = RemoteWorker("https://example.test", "secret", worker_id="laptop")
        worker.update = lambda *_args, **_kwargs: False
        uploads = []
        worker.post = lambda path, **kwargs: uploads.append((path, kwargs)) or MagicMock()
        settings = {
            "AIR_QUERY_DELAY_SECONDS": "0", "AIR_CAPTCHA_COOLDOWN_SECONDS": "0",
            "AIR_CAPTCHA_RETRIES": "2", "AIR_BATCH_REST_EVERY": "0",
        }
        with patch("air.remote_worker.GoogleAIOverviewCollector", BlockedCollector), patch.dict("os.environ", settings):
            worker.process({"run_id": "x" * 32, "filename": "queries.txt", "queries": ["blocked", "must not run"]})

        self.assertEqual(["blocked", "blocked", "blocked"], attempts)
        self.assertEqual("paused", uploads[0][1]["data"]["terminal_state"])
        self.assertEqual("1", str(uploads[0][1]["data"]["failed_count"]))

    def test_server_exposes_batch_and_combined_downloads(self):
        run_id = self.upload().get_json()["job"]["run_id"]
        self.client.post("/worker/claim", json={"worker_id": "laptop"}, headers=self.worker_headers)
        completed = self.client.post(
            f"/worker/jobs/{run_id}/complete",
            data={
                "worker_id": "laptop", "success_count": "2", "failed_count": "0",
                "workbook": (io.BytesIO(b"combined"), "amex_air_all_batches.xlsx"),
                "batches": [
                    (io.BytesIO(b"batch-one"), "amex_air_batch_1.xlsx"),
                    (io.BytesIO(b"batch-two"), "amex_air_batch_2.xlsx"),
                ],
            },
            content_type="multipart/form-data",
            headers=self.worker_headers,
        )
        self.assertEqual(200, completed.status_code)
        job = self.client.get(f"/jobs/{run_id}", headers=self.client_headers).get_json()
        self.assertEqual(2, len(job["batch_downloads"]))
        self.assertEqual(b"combined", self.client.get(job["download_url"]).data)
        self.assertEqual(b"batch-one", self.client.get(job["batch_downloads"][0]["download_url"]).data)

    def test_legacy_excel_writer_still_matches_output_contract(self):
        job_root = Path(self.temp.name) / "legacy-job"
        job_root.mkdir()
        (job_root / "job.json").write_text(json.dumps({"queries": ["one"]}), encoding="utf-8")

        class FakeCollector:
            def collect(self, query):
                return {"status": "Success", "response": query, "parsed_json": "{}", "execution_time": 0.1}

        from unittest.mock import patch
        with patch("air.worker.GoogleAIOverviewCollector", FakeCollector), patch.dict("os.environ", {"AIR_QUERY_DELAY_SECONDS": "0"}):
            run_legacy_worker(job_root)
        workbook = load_workbook(job_root / "amex_air_results.xlsx", read_only=True, data_only=True)
        self.assertEqual(
            ("Prompt", "Status", "Response", "Parsed JSON", "Top 3 Blue Links", "Execution Time (sec)"),
            tuple(cell.value for cell in workbook["Responses"][1]),
        )
        workbook.close()


if __name__ == "__main__":
    unittest.main()
