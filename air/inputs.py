from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


class InputError(ValueError):
    pass


def _clean(values):
    queries = []
    seen = set()
    for value in values:
        query = str(value or "").strip()
        if query and query.lower() != "nan" and query not in seen:
            seen.add(query)
            queries.append(query)
    if not queries:
        raise InputError("No non-empty queries were found in the uploaded file.")
    if len(queries) > 500:
        raise InputError("A single batch may contain at most 500 unique queries.")
    return queries


def read_queries(filename, content):
    suffix = Path(filename).suffix.lower()
    if suffix == ".txt":
        return _clean(content.decode("utf-8-sig").splitlines())
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.reader(StringIO(text)))
        if not rows:
            raise InputError("The CSV file is empty.")
        header = [str(value).strip().lower() for value in rows[0]]
        index = header.index("prompt") if "prompt" in header else 0
        return _clean(row[index] for row in rows[1:] if len(row) > index)
    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            workbook.close()
            raise InputError("The Excel workbook is empty.")
        names = [str(value or "").strip().lower() for value in header]
        index = names.index("prompt") if "prompt" in names else 0
        values = [row[index] for row in rows if len(row) > index]
        workbook.close()
        return _clean(values)
    raise InputError("Upload an .xlsx, .csv, or .txt file.")
