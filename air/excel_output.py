from __future__ import annotations

from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def build_results(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Responses"
    headers = ["Prompt", "Status", "Response", "Parsed JSON", "Top 3 Blue Links", "Execution Time (sec)"]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row["prompt"], row["status"], row["response"], row["parsed_json"],
            row.get("top_blue_links", "[]"), row["execution_time"]
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="006FCF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 42, "B": 12, "C": 90, "D": 100, "E": 80, "F": 20}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    if rows:
        table = Table(displayName="AirResponses", ref=f"A1:F{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    return workbook


def write_results(path, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    build_results(rows).save(path)


def results_bytes(rows):
    stream = BytesIO()
    build_results(rows).save(stream)
    return stream.getvalue()
